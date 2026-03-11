"""
Upadsfast — Google Ads Bulk Upload Spreadsheet Generator
=========================================================
Gera planilha .xlsx completa com as 5 abas de bulk upload
para qualquer produto, país e idioma.

Uso:
    python3 generate-spreadsheet.py \
        --product "CAPNOS" \
        --country "US" \
        --language "en" \
        --price "37" \
        --currency "$" \
        --discount "50" \
        --guarantee "30" \
        --ship-min "74" \
        --has-free-shipping yes \
        --url "https://seusite.com"

Saída: output/{PRODUCT}_{COUNTRY}_BulkUpload.xlsx
"""

import argparse
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)


# ─── TEMPLATES DE COPY POR IDIOMA ───────────────────────────────────────────
# Variáveis disponíveis em todos os templates:
#   {product}   → nome do produto
#   {country}   → país (US, BR, DE, etc.)
#   {discount}  → % de desconto (ex: 50)
#   {guarantee} → dias de garantia (ex: 30)
#   {price}     → preço (ex: 37)
#   {currency}  → símbolo da moeda (ex: $, R$, €)
#   {ship_min}  → valor mínimo para frete grátis (ex: 74)
#
# ATENÇÃO: todos os templates foram escritos para caber nos limites com as
# variáveis substituídas por valores CURTOS (produto ≤10 chars recomendado).
# A função trunc() garante que nada ultrapasse o limite.

COPY = {
    "en": {
        # Headlines (≤30 chars após substituição)
        "h": [
            "{product} Official Site",
            "Up To {discount}% Off Today",
            "Buy 1 Get 1 {discount}% Off",
            "Free {country} Shipping",
            "{guarantee}-Day Money-Back",
            "Order {product} For {currency}{price}",
            "Limited Time {discount}% Sale",
            "Claim Your {discount}% Off",
            "Special Offer: {discount}% Off",
            "{product} - {discount}% Off Sale",
            "Try {product} Risk-Free",
            "Free Shipping Over {currency}{ship_min}",
            "Exclusive Bundle Price",
            "Secure Your Order Now",
            "Order Today Save {discount}%",
        ],
        # Descriptions (≤90 chars após substituição)
        "d": [
            "Get up to {discount}% off {product} today. Buy 1 get 1 {discount}% off. Free {country} shipping on orders over {currency}{ship_min}.",
            "Try {product} risk-free with our {guarantee}-day money-back guarantee. Order now starting at just {currency}{price}.",
            "Limited time offer! Claim your {discount}% discount on {product} bundles today. Delivery to the {country}.",
            "Buy {product} for only {currency}{price}. Enjoy a {guarantee}-day guarantee and up to {discount}% off today.",
        ],
        # Callouts (≤25 chars)
        "callouts": [
            "Up To {discount}% Off",
            "Free Shipping Over {currency}{ship_min}",
            "{guarantee}-Day Guarantee",
            "Buy 1 Get 1 {discount}% Off",
            "Starting At {currency}{price}",
        ],
        # Sitelinks: (text ≤25, d1 ≤35, d2 ≤35)
        "sitelinks": [
            ("Claim {discount}% Off",       "Buy 1 get 1 {discount}% off today.",        "Exclusive bundle deals."),
            ("Free {country} Shipping",     "Free shipping on orders over {currency}{ship_min}.", "Fast delivery across the {country}."),
            ("{guarantee}-Day Guarantee",   "Try {product} completely risk-free.",        "100% money-back guarantee."),
            ("Order {product} Now",         "Devices starting at just {currency}{price}.", "Order from the official site."),
            ("Bundle Pricing",              "Save big with our bundles.",                 "Special discount applied."),
            ("Limited Time Sale",           "Don't miss out on {discount}% off.",         "Order today to save instantly."),
        ],
        "snippet_header": "Types",
        "snippet_values": [
            "{discount}% Discount",
            "Free {country} Shipping",
            "{guarantee}-Day Guarantee",
            "Bundles From {currency}{price}",
        ],
    },

    "pt": {
        "h": [
            "{product} Site Oficial",
            "Até {discount}% OFF Hoje",
            "Compre 1 Leve 2 c/{discount}% OFF",
            "Frete Grátis no Brasil",
            "Garantia de {guarantee} Dias",
            "{product} Por {currency}{price}",
            "Oferta Limitada {discount}% OFF",
            "Garanta {discount}% de Desconto",
            "Oferta Especial: {discount}% OFF",
            "{product} {discount}% Desconto",
            "Experimente Sem Risco",
            "Frete Grátis Acima {currency}{ship_min}",
            "Preço de Kit Exclusivo",
            "Compre com Segurança",
            "Peça Hoje, Economize {discount}%",
        ],
        "d": [
            "Até {discount}% OFF em {product} hoje. Frete grátis acima de {currency}{ship_min}. Garantia de {guarantee} dias.",
            "Experimente {product} sem risco — {guarantee} dias de garantia. Compre agora por {currency}{price}.",
            "Oferta limitada! {discount}% OFF em kits de {product} hoje. Entrega para todo o Brasil.",
            "{product} por {currency}{price}. Garantia de {guarantee} dias e até {discount}% de desconto hoje.",
        ],
        "callouts": [
            "Até {discount}% OFF",
            "Frete Grátis Acima {currency}{ship_min}",
            "Garantia {guarantee} Dias",
            "Kit com {discount}% Desconto",
            "A Partir de {currency}{price}",
        ],
        "sitelinks": [
            ("Desconto de {discount}%",     "Kit com {discount}% de desconto hoje.",      "Ofertas exclusivas de kit."),
            ("Frete Grátis",                "Frete grátis acima de {currency}{ship_min}.", "Entrega rápida no Brasil."),
            ("Garantia {guarantee} Dias",   "Experimente sem risco algum.",               "Devolução em {guarantee} dias."),
            ("Comprar {product}",           "A partir de {currency}{price} por unidade.",  "Compre no site oficial."),
            ("Preço de Kit",                "Economize mais com nossos kits.",             "Desconto aplicado automaticamente."),
            ("Oferta Tempo Limitado",       "Não perca {discount}% de desconto.",          "Peça hoje e economize."),
        ],
        "snippet_header": "Tipos",
        "snippet_values": [
            "{discount}% de Desconto",
            "Frete Grátis",
            "Garantia {guarantee} Dias",
            "Kits a Partir {currency}{price}",
        ],
    },

    "es": {
        "h": [
            "{product} Sitio Oficial",
            "Hasta {discount}% de Descuento",
            "2x1 con {discount}% OFF",
            "Envío Gratis a {country}",
            "Garantía de {guarantee} Días",
            "{product} Por Solo {currency}{price}",
            "Oferta Limitada {discount}% OFF",
            "Reclama tu {discount}% de Dto.",
            "Oferta Especial: {discount}% OFF",
            "{product} {discount}% Descuento",
            "Prueba {product} Sin Riesgo",
            "Envío Gratis Desde {currency}{ship_min}",
            "Precio Bundle Exclusivo",
            "Asegura tu Pedido Ahora",
            "Ordena Hoy, Ahorra {discount}%",
        ],
        "d": [
            "Hasta {discount}% off en {product} hoy. Envío gratis en pedidos sobre {currency}{ship_min}. Garantía {guarantee} días.",
            "Prueba {product} sin riesgo con garantía de {guarantee} días. Ordena desde {currency}{price}.",
            "¡Tiempo limitado! {discount}% de descuento en bundles de {product}. Envío a {country}.",
            "Compra {product} por {currency}{price}. Garantía {guarantee} días y hasta {discount}% de descuento hoy.",
        ],
        "callouts": [
            "Hasta {discount}% Descuento",
            "Envío Gratis Desde {currency}{ship_min}",
            "Garantía {guarantee} Días",
            "Bundle 2x1 {discount}% OFF",
            "Desde Solo {currency}{price}",
        ],
        "sitelinks": [
            ("Descuento {discount}%",       "Bundle con {discount}% off hoy.",            "Ofertas exclusivas de bundle."),
            ("Envío Gratis",                "Envío gratis en pedidos sobre {currency}{ship_min}.", "Entrega rápida a {country}."),
            ("Garantía {guarantee} Días",   "Prueba {product} completamente gratis.",      "Reembolso 100% garantizado."),
            ("Ordenar {product} Ya",        "Dispositivos desde solo {currency}{price}.",  "Compra en el sitio oficial."),
            ("Precios Bundle",              "Ahorra más con nuestros bundles.",            "Descuento aplicado al instante."),
            ("Oferta Tiempo Limitado",      "No pierdas el {discount}% off.",              "Ordena hoy y ahorra al instante."),
        ],
        "snippet_header": "Types",
        "snippet_values": [
            "{discount}% Descuento",
            "Envío Gratis",
            "Garantía {guarantee} Días",
            "Bundles Desde {currency}{price}",
        ],
    },

    "de": {
        "h": [
            "{product} Offizieller Shop",
            "Bis zu {discount}% Rabatt Heute",
            "2 Kaufen, 1 mit {discount}% Off",
            "Kostenloser Versand",
            "{guarantee} Tage Geld-zurück",
            "{product} Ab {currency}{price}",
            "Zeitlich begrenzt {discount}% Off",
            "{discount}% Rabatt Jetzt Sichern",
            "Sonderangebot: {discount}% Rabatt",
            "{product} {discount}% Sparen",
            "{product} Risikofrei Testen",
            "Gratis Versand ab {currency}{ship_min}",
            "Exklusiver Bundle-Preis",
            "Jetzt Sicher Bestellen",
            "Heute Bestellen, {discount}% Sparen",
        ],
        "d": [
            "Bis zu {discount}% auf {product}. Gratis Versand ab {currency}{ship_min}. {guarantee} Tage Garantie.",
            "{product} risikofrei testen – {guarantee} Tage Geld-zurück. Jetzt ab {currency}{price} bestellen.",
            "Zeitlich begrenzt! {discount}% Rabatt auf {product} Bundles. Lieferung nach {country}.",
            "{product} für {currency}{price}. {guarantee} Tage Garantie und bis zu {discount}% Rabatt heute.",
        ],
        "callouts": [
            "Bis {discount}% Rabatt",
            "Gratis Versand ab {currency}{ship_min}",
            "{guarantee} Tage Garantie",
            "Bundle {discount}% Rabatt",
            "Ab {currency}{price}",
        ],
        "sitelinks": [
            ("{discount}% Rabatt Sichern",   "Bundle mit {discount}% Rabatt heute.",       "Exklusive Bundle-Angebote."),
            ("Kostenloser Versand",           "Gratis Versand ab {currency}{ship_min}.",    "Schnelle Lieferung nach {country}."),
            ("{guarantee}-Tage-Garantie",     "{product} komplett risikofrei testen.",      "100% Geld-zurück-Garantie."),
            ("{product} Jetzt Bestellen",     "Geräte ab {currency}{price}.",               "Im offiziellen Shop kaufen."),
            ("Bundle-Preise",                 "Mit Bundles mehr sparen.",                   "Rabatt automatisch abgezogen."),
            ("Zeitlich Begrenzt",             "{discount}% Rabatt nicht verpassen.",         "Heute bestellen und sparen."),
        ],
        "snippet_header": "Types",
        "snippet_values": [
            "{discount}% Rabatt",
            "Kostenloser Versand",
            "{guarantee} Tage Garantie",
            "Bundles ab {currency}{price}",
        ],
    },

    "da": {
        "h": [
            "{product} Officielt Site",
            "Op til {discount}% Rabat I Dag",
            "60-Dages Pengene-Tilbage",
            "Hurtig Levering til EU",
            "Bestil Nu fra {currency}{price}",
            "Tidsbegrænset {discount}% Tilbud",
            "Spar {discount}% på {product}",
            "Særtilbud: {discount}% Rabat",
            "{product} - {discount}% Rabat",
            "Prøv {product} Risikofrit",
            "Gratis EU Levering",
            "{guarantee}-Dages Tilfredshed",
            "Eksklusiv Bundle Pris",
            "Sikker & Diskret Kasse",
            "Bestil I Dag, Spar {discount}%",
        ],
        "d": [
            "Op til {discount}% rabat på {product} i dag. 60-dages pengene-tilbage garanti. Hurtig EU levering.",
            "Prøv {product} risikofrit med vores {guarantee}-dages tilfredshedsgaranti. Bestil fra {currency}{price}.",
            "Tidsbegrænset tilbud! {discount}% rabat på {product} i dag. Hurtig og diskret levering til {country}.",
            "Køb {product} for kun {currency}{price}. Nyd {guarantee}-dages garanti og op til {discount}% rabat i dag.",
        ],
        "callouts": [
            "Op til {discount}% Rabat",
            "Gratis EU Levering",
            "{guarantee}-Dages Garanti",
            "Sikker & Diskret Kasse",
            "Fra {currency}{price}",
        ],
        "sitelinks": [
            ("Spar {discount}% I Dag",        "Op til {discount}% rabat på {product}.",     "Eksklusivt bundle tilbud."),
            ("Gratis EU Levering",             "Hurtig levering til hele EU.",               "Diskret og sikker forsendelse."),
            ("{guarantee}-Dages Garanti",      "Prøv {product} helt risikofrit.",            "Fuld pengene-tilbage garanti."),
            ("Bestil {product} Nu",            "Kapsler fra kun {currency}{price}.",         "Bestil fra det officielle site."),
            ("Bundle Pris",                    "Spar mere med vores bundles.",               "Rabat aktiveres automatisk."),
            ("Tidsbegrænset Tilbud",           "Gå ikke glip af {discount}% rabat.",         "Bestil i dag og spar straks."),
        ],
        "snippet_header": "Types",
        "snippet_values": [
            "{discount}% Rabat",
            "Gratis EU Levering",
            "{guarantee}-Dages Garanti",
            "Fra {currency}{price}",
        ],
    },
}


# ─── HELPERS ────────────────────────────────────────────────────────────────

def render(template: str, ctx: dict) -> str:
    """Substitui variáveis no template."""
    return template.format(**ctx)


def trunc(text: str, limit: int) -> str:
    """Trunca o texto no limite de caracteres, preservando palavras inteiras."""
    if len(text) <= limit:
        return text
    truncated = text[:limit]
    last_space = truncated.rfind(" ")
    return truncated[:last_space].rstrip() if last_space > 0 else truncated


def render_trunc(template: str, ctx: dict, limit: int) -> str:
    return trunc(render(template, ctx), limit)


def validate(label: str, text: str, limit: int, errors: list) -> str:
    """Valida o limite e registra erro se necessário."""
    if len(text) > limit:
        errors.append(f"{label} EXCEDE {limit} chars [{len(text)}]: '{text}'")
    return text


# ─── ESTILOS ────────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_AUX    = PatternFill("solid", fgColor="2E75B6")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_DATA   = Font(size=10)
ALIGN_CTR   = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT  = Alignment(horizontal="left",   vertical="center")
_SIDE       = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

TAB_COLORS = {
    "Campaigns":           "0B5394",
    "Ad groups":           "274E13",
    "Ads":                 "1F4E79",
    "Callouts":            "375623",
    "Sitelinks":           "7B2C2C",
    "Snippets":            "614A19",
    "Promotions":          "4A235A",
}


def hdr(ws, row, col, value, aux=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = (
        FILL_AUX if aux else FILL_HEADER, FONT_HEADER, ALIGN_CTR, BORDER)
    return c


def dat(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.alignment, c.border = FONT_DATA, ALIGN_LEFT, BORDER
    return c


def auto_width(ws, mn=12, mx=45):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)


# ─── GERADOR PRINCIPAL ──────────────────────────────────────────────────────

def generate(cfg: dict) -> str:
    """
    Gera o arquivo .xlsx e retorna o caminho do arquivo gerado.

    cfg keys obrigatórias:
        product, country, language, price, currency,
        discount, guarantee, ship_min, has_free_shipping, url
    """
    lang = cfg["language"].lower()
    if lang not in COPY:
        print(f"⚠️  Idioma '{lang}' não encontrado. Usando 'en' como fallback.")
        lang = "en"

    tmpl = COPY[lang]
    ctx  = {
        "product":   cfg["product"],
        "country":   cfg["country"],
        "discount":  cfg["discount"],
        "guarantee": cfg["guarantee"],
        "price":     cfg["price"],
        "currency":  cfg["currency"],
        "ship_min":  cfg["ship_min"],
    }

    campaign = f"Search - {cfg['product']} - {cfg['country']}"
    ad_group = f"{cfg['product']} - Offer"
    path1    = trunc(cfg["product"], 15)
    path2    = trunc(f"{cfg['discount']}-Off", 15)
    url      = cfg["url"]

    errors = []

    # ── Headlines
    headlines = []
    for i, t in enumerate(tmpl["h"][:15], 1):
        text = render_trunc(t, ctx, 30)
        validate(f"H{i:02d}", text, 30, errors)
        headlines.append(text)

    # ── Descriptions
    descriptions = []
    for i, t in enumerate(tmpl["d"][:4], 1):
        text = render_trunc(t, ctx, 90)
        validate(f"D{i}", text, 90, errors)
        descriptions.append(text)

    # ── Callouts
    callouts = []
    for t in tmpl["callouts"]:
        text = render_trunc(t, ctx, 25)
        validate(f"Callout", text, 25, errors)
        callouts.append(text)

    # ── Sitelinks
    sitelinks = []
    for txt_t, d1_t, d2_t in tmpl["sitelinks"]:
        txt = render_trunc(txt_t, ctx, 25)
        d1  = render_trunc(d1_t,  ctx, 35)
        d2  = render_trunc(d2_t,  ctx, 35)
        validate(f"Sitelink text '{txt}'", txt, 25, errors)
        validate(f"Sitelink D1 '{txt}'",   d1,  35, errors)
        validate(f"Sitelink D2 '{txt}'",   d2,  35, errors)
        sitelinks.append((txt, url, d1, d2))

    # ── Snippet
    snip_header = tmpl["snippet_header"]
    snip_values = [render_trunc(v, ctx, 25) for v in tmpl["snippet_values"][:4]]

    # ── Promoção
    promo = ("None", "Percent", int(cfg["discount"]), "", url, "", "")

    # ── Path validation
    validate("Path 1", path1, 15, errors)
    validate("Path 2", path2, 15, errors)

    # ── Snippet values concatenados (separados por ;)
    snip_values_str = ";".join(snip_values)

    # ════════════════════════════════════════════════════════════
    # BUILD WORKBOOK
    # ════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── ABA 1: Campaigns (criar campanha nova)
    ws0 = wb.create_sheet("Campaigns")
    ws0.sheet_properties.tabColor = TAB_COLORS["Campaigns"]
    ws0.row_dimensions[1].height = 22
    camp_headers = [
        "Campaign", "Campaign type", "Networks", "Budget",
        "Bid Strategy type", "EU political ads", "Status",
    ]
    for c, h in enumerate(camp_headers, 1):
        hdr(ws0, 1, c, h)
    camp_data = [
        campaign, "Search", "Google Search", cfg.get("budget", "50"),
        "Maximize clicks", "No", "Enabled",
    ]
    for c, v in enumerate(camp_data, 1):
        dat(ws0, 2, c, v)
    ws0.freeze_panes = "A2"
    auto_width(ws0)

    # ── ABA 2: Ad groups
    ws_ag = wb.create_sheet("Ad groups")
    ws_ag.sheet_properties.tabColor = TAB_COLORS["Ad groups"]
    ws_ag.row_dimensions[1].height = 22
    for c, h in enumerate(["Campaign", "Ad group", "Status"], 1):
        hdr(ws_ag, 1, c, h)
    dat(ws_ag, 2, 1, campaign)
    dat(ws_ag, 2, 2, ad_group)
    dat(ws_ag, 2, 3, "Enabled")
    ws_ag.freeze_panes = "A2"
    auto_width(ws_ag)

    # ── ABA 3: Ads (RSA)
    ws1 = wb.create_sheet("Ads")
    ws1.sheet_properties.tabColor = TAB_COLORS["Ads"]
    ws1.row_dimensions[1].height  = 22

    official = (
        ["Campaign", "Ad group", "Final URL"]
        + [f"Headline {i}" for i in range(1, 16)]
        + [f"Description {i}" for i in range(1, 5)]
        + ["Path 1", "Path 2", "Status"]
    )
    auxiliary = [
        "Product_Name", "Country", "Language",
        "Discount_Value", "Guarantee_Days",
        "Has_Free_Shipping", "Free_Ship_Min",
        "Currency", "Price",
    ]

    for c, h in enumerate(official, 1):
        hdr(ws1, 1, c, h)
    aux_start = len(official) + 1
    for i, h in enumerate(auxiliary):
        hdr(ws1, 1, aux_start + i, h, aux=True)

    row_data = (
        [campaign, ad_group, url]
        + headlines
        + descriptions
        + [path1, path2, "Enabled"]
    )
    for c, v in enumerate(row_data, 1):
        dat(ws1, 2, c, v)

    aux_data = [
        cfg["product"], cfg["country"], cfg["language"],
        cfg["discount"], cfg["guarantee"],
        cfg["has_free_shipping"], cfg["ship_min"],
        cfg["currency"], cfg["price"],
    ]
    for i, v in enumerate(aux_data):
        dat(ws1, 2, aux_start + i, v)

    ws1.freeze_panes = "A2"
    auto_width(ws1)

    # ── ABA 4: Callouts
    ws2 = wb.create_sheet("Callouts")
    ws2.sheet_properties.tabColor = TAB_COLORS["Callouts"]
    ws2.row_dimensions[1].height  = 22
    for c, h in enumerate(["Campaign", "Action", "Callout text", "Status"], 1):
        hdr(ws2, 1, c, h)
    for r, text in enumerate(callouts, 2):
        dat(ws2, r, 1, campaign)
        dat(ws2, r, 2, "Add")
        dat(ws2, r, 3, text)
        dat(ws2, r, 4, "Enabled")
    ws2.freeze_panes = "A2"
    auto_width(ws2)

    # ── ABA 5: Sitelinks
    ws3 = wb.create_sheet("Sitelinks")
    ws3.sheet_properties.tabColor = TAB_COLORS["Sitelinks"]
    ws3.row_dimensions[1].height  = 22
    for c, h in enumerate(["Campaign", "Action", "Sitelink text", "Final URL", "Description line 1", "Description line 2", "Status"], 1):
        hdr(ws3, 1, c, h)
    for r, (txt, lnk, d1, d2) in enumerate(sitelinks, 2):
        dat(ws3, r, 1, campaign)
        dat(ws3, r, 2, "Add")
        dat(ws3, r, 3, txt)
        dat(ws3, r, 4, lnk)
        dat(ws3, r, 5, d1)
        dat(ws3, r, 6, d2)
        dat(ws3, r, 7, "Enabled")
    ws3.freeze_panes = "A2"
    auto_width(ws3)

    # ── ABA 6: Snippets
    ws4 = wb.create_sheet("Snippets")
    ws4.sheet_properties.tabColor = TAB_COLORS["Snippets"]
    ws4.row_dimensions[1].height  = 22
    for c, h in enumerate(["Campaign", "Action", "Structured snippet header", "Structured snippet values", "Status"], 1):
        hdr(ws4, 1, c, h)
    dat(ws4, 2, 1, campaign)
    dat(ws4, 2, 2, "Add")
    dat(ws4, 2, 3, snip_header)
    dat(ws4, 2, 4, snip_values_str)
    dat(ws4, 2, 5, "Enabled")
    ws4.freeze_panes = "A2"
    auto_width(ws4)

    # ── ABA 7: Promotions
    ws5 = wb.create_sheet("Promotions")
    ws5.sheet_properties.tabColor = TAB_COLORS["Promotions"]
    ws5.row_dimensions[1].height  = 22
    for c, h in enumerate(["Campaign", "Action", "Occasion", "Discount type", "Percent off", "Promotion code", "Final URL", "Start date", "End date", "Status"], 1):
        hdr(ws5, 1, c, h)
    promo_data = [
        campaign, "Add", "None", "Percent", int(cfg["discount"]),
        "", url, "", "", "Enabled",
    ]
    for c, v in enumerate(promo_data, 1):
        dat(ws5, 2, c, v)
    ws5.freeze_panes = "A2"
    auto_width(ws5)

    # ── SALVAR
    output_dir = Path(__file__).parent.parent / "output"
    output_dir.mkdir(exist_ok=True)
    filename = f"{cfg['product'].upper()}_{cfg['country'].upper()}_BulkUpload.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)

    # ── RELATÓRIO
    print(f"\n{'='*55}")
    print(f"  UPADSFAST — Bulk Upload Generator")
    print(f"{'='*55}")
    print(f"  Produto   : {cfg['product']}")
    print(f"  País      : {cfg['country']}  |  Idioma: {cfg['language']}")
    print(f"  Campanha  : {campaign}")
    print(f"  Budget    : {cfg.get('budget', '50')}")
    print(f"  EU Ads    : No")
    print(f"{'='*55}\n")

    print("── Campanha")
    print(f"  Campaign type     : Search")
    print(f"  Budget            : {cfg.get('budget', '50')}")
    print(f"  Bid Strategy type : Maximize clicks")
    print(f"  EU political ads  : No")

    print("\n── Headlines")
    for i, h in enumerate(headlines, 1):
        print(f"  H{i:02d} [{len(h):2d}/30] ✅  {h}")

    print("\n── Descriptions")
    for i, d in enumerate(descriptions, 1):
        print(f"  D{i}  [{len(d):2d}/90] ✅  {d}")

    print("\n── Callouts (Action: Add)")
    for c in callouts:
        print(f"  [{len(c):2d}/25] ✅  {c}")

    print("\n── Sitelinks (Action: Add)")
    for txt, _, d1, d2 in sitelinks:
        print(f"  text [{len(txt):2d}/25] ✅  {txt}")
        print(f"    D1 [{len(d1):2d}/35] ✅  {d1}")
        print(f"    D2 [{len(d2):2d}/35] ✅  {d2}")

    print(f"\n── Snippet (Action: Add): {snip_header} → {snip_values_str}")
    print(f"── Promoção (Action: Add): {cfg['discount']}% off")
    print(f"── Path 1: {path1}  |  Path 2: {path2}")

    if errors:
        print(f"\n❌ {len(errors)} ERRO(S) ENCONTRADO(S):")
        for e in errors:
            print(f"  ⚠️  {e}")
    else:
        print(f"\n✅ Todos os limites respeitados.")

    print(f"\n📄 Arquivo: {output_path}\n")
    return str(output_path)


# ─── CLI ────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Upadsfast — Gera planilha de bulk upload para Google Ads."
    )
    p.add_argument("--product",           required=True,  help="Nome do produto (ex: CAPNOS)")
    p.add_argument("--country",           required=True,  help="País (ex: US, BR, DE)")
    p.add_argument("--language",          required=True,  help="Idioma: en | pt | es | de")
    p.add_argument("--price",             required=True,  help="Preço (ex: 37)")
    p.add_argument("--currency",          default="$",    help="Símbolo da moeda (ex: $, R$, €)")
    p.add_argument("--discount",          required=True,  help="% de desconto (ex: 50)")
    p.add_argument("--guarantee",         required=True,  help="Dias de garantia (ex: 30)")
    p.add_argument("--ship-min",          default="0",    help="Valor mínimo frete grátis (ex: 74)")
    p.add_argument("--has-free-shipping", default="Yes",  help="Frete grátis? Yes | No")
    p.add_argument("--budget",            default="50",   help="Orçamento diário da campanha (ex: 50)")
    p.add_argument("--url",               required=True,  help="URL final do anúncio")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    cfg = {
        "product":          args.product,
        "country":          args.country,
        "language":         args.language,
        "price":            args.price,
        "currency":         args.currency,
        "discount":         args.discount,
        "guarantee":        args.guarantee,
        "ship_min":         args.ship_min,
        "has_free_shipping": args.has_free_shipping,
        "budget":           args.budget,
        "url":              args.url,
    }
    generate(cfg)
